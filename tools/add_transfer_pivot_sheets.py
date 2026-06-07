from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FILES = [
    Path(r"C:\Users\BCIMIT\Downloads\FW_ List of Assets at Site\Material tranfer Hyderabad to Bangalore.xlsx"),
    Path(r"C:\Users\BCIMIT\Downloads\FW_ List of Assets at Site\MATERIAL TRANSFER DETAILS HYD TO MUMBAI.xlsx"),
    Path(r"C:\Users\BCIMIT\Downloads\FW_ List of Assets at Site\Material Transfer details.xlsx"),
    Path(r"C:\Users\BCIMIT\Downloads\FW_ List of Assets at Site\RMD MTO Detail HYD TO MUMBAI.xlsx"),
]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def row_values(ws, row_number: int) -> list:
    return [cell.value for cell in ws[row_number]]


def find_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    max_row = min(ws.max_row, 30)
    for row_number in range(1, max_row + 1):
        values = row_values(ws, row_number)
        non_empty = sum(1 for value in values if clean(value))
        alpha = sum(1 for value in values if re.search(r"[A-Za-z]", clean(value)))
        score = non_empty + alpha
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row


def unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    output = []
    for index, header in enumerate(headers, start=1):
        name = header or f"Column {index}"
        seen[name] = seen.get(name, 0) + 1
        output.append(name if seen[name] == 1 else f"{name} {seen[name]}")
    return output


def useful_max_col(ws, scan_rows: int = 60) -> int:
    max_col = 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_rows)):
        for cell in row:
            if clean(cell.value):
                max_col = max(max_col, cell.column)
    return max_col


def is_identifier_header(header: str) -> bool:
    return bool(re.search(r"gate pass\s*(number|no\.?$)|sl\s*no|vehicle|docket\s*(number|no\.?$)", header, re.I))


def is_transfer_header(header: str) -> bool:
    return bool(re.search(r"gate pass|docket|transferred qty", header, re.I)) and not is_identifier_header(header)


def is_quantity_header(header: str) -> bool:
    return bool(re.search(r"\bqty\b|quantity|approved|jll|yet|balance|stock|requirement|excess|transfer order", header, re.I))


def detect_columns(ws):
    header_row = find_header_row(ws)
    max_col = useful_max_col(ws)
    headers = unique_headers([clean(ws.cell(header_row, col).value) for col in range(1, max_col + 1)])

    material_col = None
    unit_col = None
    for col, header in enumerate(headers, start=1):
        if material_col is None and re.search(r"material|description|item", header, re.I):
            material_col = col
        if unit_col is None and re.search(r"\buom\b|unit", header, re.I):
            unit_col = col

    # Special merged-heading Mumbai format: gate-pass names are in row 5.
    gatepass_row = header_row + 1
    gatepass_cols = []
    if gatepass_row <= ws.max_row:
        for col in range(1, max_col + 1):
            label = clean(ws.cell(gatepass_row, col).value)
            if re.search(r"gate pass|docket", label, re.I):
                gatepass_cols.append((col, label))

    numeric_cols = []
    total_cols = []
    for col, header in enumerate(headers, start=1):
        if col in (material_col, unit_col):
            continue
        if is_identifier_header(header):
            continue
        wanted_header = is_quantity_header(header) or is_transfer_header(header)
        numeric_count = 0
        for row in range(header_row + 1, min(ws.max_row, header_row + 35) + 1):
            if as_number(ws.cell(row, col).value) is not None:
                numeric_count += 1
        if wanted_header or numeric_count >= 2:
            numeric_cols.append((col, header))
            if is_transfer_header(header) or re.search(r"^qty$", header, re.I):
                total_cols.append(len(numeric_cols) - 1)

    if gatepass_cols:
        numeric_cols = gatepass_cols
        total_cols = list(range(len(numeric_cols)))
        data_start_row = gatepass_row + 3
    else:
        data_start_row = header_row + 1

    if not total_cols:
        total_cols = [
            index
            for index, (_, header) in enumerate(numeric_cols)
            if re.search(r"total transfer|transferred|^qty$", header, re.I)
        ]

    return header_row, data_start_row, material_col, unit_col, numeric_cols, total_cols


def build_pivot_rows(ws):
    header_row, data_start_row, material_col, unit_col, numeric_cols, total_cols = detect_columns(ws)
    if not material_col or not numeric_cols:
        return None

    pivot: dict[tuple[str, str], list[float]] = {}
    for row in range(data_start_row, ws.max_row + 1):
        material = clean(ws.cell(row, material_col).value)
        if not material or material.startswith("#") or material.lower().startswith(("approved", "total")):
            continue
        unit = clean(ws.cell(row, unit_col).value) if unit_col else ""
        values = []
        has_value = False
        for col, _ in numeric_cols:
            number = as_number(ws.cell(row, col).value)
            if number is None:
                values.append(0.0)
            else:
                values.append(number)
                has_value = True
        if not has_value:
            continue
        key = (material, unit)
        if key not in pivot:
            pivot[key] = [0.0] * len(numeric_cols)
        for index, value in enumerate(values):
            pivot[key][index] += value

    if not pivot:
        return None

    headers = ["Material Description", "Unit"] + [name for _, name in numeric_cols]
    if total_cols:
        headers.append("Transfer / Qty Total")
    rows = []
    for (material, unit), values in sorted(pivot.items(), key=lambda item: item[0][0].lower()):
        row = [material, unit] + values
        if total_cols:
            row.append(sum(values[index] for index in total_cols))
        rows.append(row)
    total_row = ["Grand Total", ""] + [sum(row[index + 2] for row in rows) for index in range(len(numeric_cols))]
    if total_cols:
        total_row.append(sum(row[-1] for row in rows))
    rows.append(total_row)
    return headers, rows


def safe_sheet_name(wb, base: str) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", base)[:31]
    if name not in wb.sheetnames:
        return name
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{name[:31 - len(suffix)]}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
        index += 1


def remove_old_pivots(wb):
    for sheet_name in list(wb.sheetnames):
        if sheet_name.startswith("Pivot_"):
            del wb[sheet_name]


def format_pivot_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = total_fill
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(ws.max_row, 60) + 1):
            max_len = max(max_len, len(clean(ws.cell(row, col).value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 36)
        if col >= 3:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col).number_format = "#,##0.##"


def add_pivots(path: Path):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
    shutil.copy2(path, backup_path)

    wb = openpyxl.load_workbook(path)
    source_sheets = [sheet for sheet in wb.worksheets if not sheet.title.startswith("Pivot_")]
    remove_old_pivots(wb)

    created = []
    for source in source_sheets:
        pivot_data = build_pivot_rows(source)
        if not pivot_data:
            continue
        headers, rows = pivot_data
        pivot_name = safe_sheet_name(wb, f"Pivot_{source.title}")
        pivot_ws = wb.create_sheet(pivot_name)
        pivot_ws.append(headers)
        for row in rows:
            pivot_ws.append(row)
        format_pivot_sheet(pivot_ws)
        created.append(pivot_name)

    wb.save(path)
    return backup_path, created


def main():
    for path in FILES:
        backup_path, created = add_pivots(path)
        print(path.name)
        print(f"  backup: {backup_path}")
        print(f"  pivot sheets: {', '.join(created) if created else 'none'}")


if __name__ == "__main__":
    main()
