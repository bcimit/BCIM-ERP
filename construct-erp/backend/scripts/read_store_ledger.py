import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\BCIMIT\Desktop\Store ledger.xlsx", data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n{'='*60}\nSHEET: {sheet}  (rows={ws.max_row} cols={ws.max_column})\n{'='*60}")
    for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
        vals = [str(c.value).strip() if c.value is not None else "" for c in row]
        non_empty = [v for v in vals if v]
        if non_empty:
            print(" | ".join(vals[:10]))
