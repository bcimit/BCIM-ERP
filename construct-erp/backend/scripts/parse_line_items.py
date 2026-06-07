"""
Parses all 9 LANCO PO/WO Excel files and generates SQL INSERT statements
for po_items and work_order_items.
"""
import openpyxl, re, textwrap
from pathlib import Path

BASE = Path(r"H:\OFFICE PROJECTS\consrpro\construct-erp-local\construct-erp\backend\uploads\documents")

# Maps order number → (file, type, db_id)
ORDERS = {
    "POLANLH10001": ("ab6d6cbc-7cab-457e-8f80-e745f55a8245.xlsx", "PO"),
    "POLANLH10002": ("94a51a26-50ff-4f4f-ad5f-a4eb0882d237.xlsx", "PO"),
    "POLANLH10003": ("2e158ead-d641-43c3-8830-acc8b0b44116.xlsx", "PO"),
    "POLANLH10004": ("1c692162-3d5a-4e97-9df5-14001921a462.xlsx", "PO"),
    "WOLANLH10001": ("54fd83bb-5275-4fc4-8f28-30321548f465.xlsx", "WO"),
    "WOLANLH10002": ("f6ccebac-d8c9-45a5-a3b7-c79202cb43f3.xlsx", "WO"),
    "WOLANLH10003": ("a509c8f2-8f6d-4280-900b-1837153b234b.xlsx", "WO"),
    "WOLANLH10004": ("dc72c6d9-2f40-4fd4-8fda-b48ae03151ad.xlsx", "WO"),
    "WOLANLH10005": ("ad2a805f-d79d-44c9-8c3a-433d28a4a788.xlsx", "WO"),
}

def cell_val(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ").replace("'", "''")

def to_num(v):
    """Convert cell value to float, return None if not numeric."""
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except:
        return None

def escape(s):
    return s.replace("'", "''")

# ── PO item parser ────────────────────────────────────────────────────────────
def parse_po(ws, order_no):
    """Returns list of dicts with PO line item fields."""
    rows = list(ws.iter_rows())
    # Find header row (contains "Sl No" or "Sl" and "Description")
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(c.value or "").lower() for c in row]
        joined = " ".join(vals)
        if "description" in joined and ("qty" in joined or "quantity" in joined or "uom" in joined):
            header_row = i
            break
    if header_row is None:
        print(f"  [WARN] {order_no}: header row not found")
        return []

    # Map column indices from header
    hdr = rows[header_row]
    col_idx = {}
    for j, c in enumerate(hdr):
        v = str(c.value or "").lower().strip()
        if "description" in v:       col_idx["desc"] = j
        elif "uom" in v or "unit" in v: col_idx["unit"] = j
        elif "qty" in v or "quantity" in v: col_idx["qty"] = j
        elif "rate" in v:            col_idx["rate"] = j
        elif "amount" in v:          col_idx["amount"] = j
        elif "hsn" in v:             col_idx["hsn"] = j
        elif "sl" in v or "no" == v: col_idx["sl"] = j

    items = []
    for row in rows[header_row + 1:]:
        # Stop at sub total / grand total row
        first_val = str(row[0].value or "").lower()
        if any(k in first_val for k in ["sub total", "subtotal", "grand total", "grandtotal", "total"]):
            break
        if all(c.value is None for c in row):
            continue

        desc = cell_val(row[col_idx.get("desc", 1)]) if col_idx.get("desc") is not None else ""
        unit = cell_val(row[col_idx.get("unit", 2)]) if col_idx.get("unit") is not None else ""
        qty  = to_num(row[col_idx.get("qty",  3)].value) if col_idx.get("qty")  is not None else None
        rate = to_num(row[col_idx.get("rate", 4)].value) if col_idx.get("rate") is not None else None
        amt  = to_num(row[col_idx.get("amount", 5)].value) if col_idx.get("amount") is not None else None
        hsn  = cell_val(row[col_idx.get("hsn", 99)]) if col_idx.get("hsn") is not None else ""

        # skip if no meaningful desc and no rate
        if not desc and rate is None:
            continue
        if qty is None and rate is None:
            continue

        # Infer amount
        if amt is None and qty is not None and rate is not None:
            amt = round(qty * rate, 2)
        if qty is None: qty = 0
        if rate is None: rate = 0
        if amt is None: amt = 0

        items.append({
            "desc": desc, "unit": unit or "Nos",
            "qty": qty, "rate": rate, "amount": amt, "hsn": hsn
        })

    return items

# ── WO item parser ────────────────────────────────────────────────────────────
def parse_wo(ws, order_no):
    """Returns list of dicts with WO line item fields."""
    rows = list(ws.iter_rows())
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(c.value or "").lower() for c in row]
        joined = " ".join(vals)
        if "description" in joined and "rate" in joined:
            header_row = i
            break
    if header_row is None:
        print(f"  [WARN] {order_no}: header row not found")
        return []

    hdr = rows[header_row]
    col_idx = {}
    for j, c in enumerate(hdr):
        v = str(c.value or "").lower().strip()
        if "description" in v or "particular" in v or "work" in v:
            if "desc" not in col_idx: col_idx["desc"] = j
        elif "uom" in v or "unit" in v: col_idx["unit"] = j
        elif "qty" in v or "quantity" in v or "nos" in v: col_idx["qty"] = j
        elif "rate" in v:            col_idx["rate"] = j
        elif "amount" in v or "total" in v: col_idx["amount"] = j

    items = []
    for row in rows[header_row + 1:]:
        first_val = str(row[0].value or "").lower()
        if any(k in first_val for k in ["sub total", "subtotal", "grand total", "grandtotal",
                                         "basic amount", "cgst", "sgst", "igst", "net total",
                                         "total amount", "gst", "tax"]):
            break
        if all(c.value is None for c in row):
            continue

        desc = cell_val(row[col_idx.get("desc", 1)]) if col_idx.get("desc") is not None else ""
        unit = cell_val(row[col_idx.get("unit", 2)]) if col_idx.get("unit") is not None else ""
        qty  = to_num(row[col_idx.get("qty",  3)].value) if col_idx.get("qty")  is not None else None
        rate = to_num(row[col_idx.get("rate", 4)].value) if col_idx.get("rate") is not None else None
        amt  = to_num(row[col_idx.get("amount", 5)].value) if col_idx.get("amount") is not None else None

        if not desc and rate is None:
            continue
        if qty is None and rate is None:
            continue

        if amt is None and qty is not None and rate is not None:
            amt = round(qty * rate, 2)
        if qty is None: qty = 0
        if rate is None: rate = 0
        if amt is None: amt = 0

        items.append({
            "desc": desc, "unit": unit or "LS",
            "qty": qty, "rate": rate, "amount": amt
        })

    return items

# ── Main ──────────────────────────────────────────────────────────────────────
print("BEGIN;\n")
print("-- ============================================================")
print("-- LANCO Hills LH 10 — Line Items for POs and WOs")
print("-- ============================================================\n")

for order_no, (fname, otype) in ORDERS.items():
    path = BASE / fname
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    if otype == "PO":
        items = parse_po(ws, order_no)
        if not items:
            print(f"-- WARNING: No items parsed for {order_no}")
            continue
        print(f"-- ── {order_no} ({len(items)} items) ──────────")
        print(f"INSERT INTO po_items (po_id, material_name, hsn_code, quantity, unit, rate, gst_rate, gst_amount, total_amount, sort_order)")
        print(f"SELECT po.id,")
        vals = []
        for i, item in enumerate(items, 1):
            # Default GST 18%, will note 5% items via comment
            gst_r = 18.00
            gst_a = round(item["amount"] * gst_r / 100, 2)
            total = round(item["amount"] + gst_a, 2)
            hsn = f"'{item['hsn']}'" if item["hsn"] else "NULL"
            v = (f"  ('{escape(item['desc'])}', {hsn}, "
                 f"{item['qty']}, '{item['unit']}', {item['rate']}, "
                 f"{gst_r}, {gst_a}, {total}, {i})")
            vals.append(v)
        print(",\n".join(f"  v.material_name, v.hsn_code, v.quantity, v.unit, v.rate, v.gst_rate, v.gst_amount, v.total_amount, v.sort_order" if j==0 else "" for j,_ in enumerate(vals)))

        # Use VALUES with cross join
        print(f"INSERT INTO po_items (po_id, material_name, hsn_code, quantity, unit, rate, gst_rate, gst_amount, total_amount, sort_order)")
        print(f"SELECT po.id, v.* FROM purchase_orders po,")
        print(f"(VALUES")
        rows_sql = []
        for i, item in enumerate(items, 1):
            gst_r = 18.00
            gst_a = round(item["amount"] * gst_r / 100, 2)
            total = round(item["amount"] + gst_a, 2)
            hsn = f"'{item['hsn']}'" if item["hsn"] else "NULL::varchar"
            row = (f"  ('{escape(item['desc'])}', {hsn}, "
                   f"{item['qty']}, '{item['unit']}', {item['rate']}, "
                   f"{gst_r}, {gst_a}, {total}, {i})")
            rows_sql.append(row)
        print(",\n".join(rows_sql))
        print(f") AS v(material_name, hsn_code, quantity, unit, rate, gst_rate, gst_amount, total_amount, sort_order)")
        print(f"WHERE po.po_number = '{order_no}'")
        print(f"ON CONFLICT DO NOTHING;\n")

    else:  # WO
        items = parse_wo(ws, order_no)
        if not items:
            print(f"-- WARNING: No items parsed for {order_no}")
            continue
        print(f"-- ── {order_no} ({len(items)} items) ──────────")
        print(f"INSERT INTO work_order_items (wo_id, description, unit, quantity, rate, remarks)")
        print(f"SELECT wo.id, v.* FROM work_orders wo,")
        print(f"(VALUES")
        rows_sql = []
        for item in items:
            row = (f"  ('{escape(item['desc'])}', '{item['unit']}', "
                   f"{item['qty']}, {item['rate']}, NULL::text)")
            rows_sql.append(row)
        print(",\n".join(rows_sql))
        print(f") AS v(description, unit, quantity, rate, remarks)")
        print(f"WHERE wo.wo_number = '{order_no}'")
        print(f"ON CONFLICT DO NOTHING;\n")

print("COMMIT;")
