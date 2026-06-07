"""Read the sample PO Excel and print all cell contents with row/col positions."""
import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\BCIMIT\Downloads\POLANLH10002- Royal Electricals.xlsx", data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}\nSHEET: {sheet_name}\n{'='*60}")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"  [{cell.coordinate}] = {repr(cell.value)}")
